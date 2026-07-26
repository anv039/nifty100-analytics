import openpyxl
from pathlib import Path
wb = openpyxl.load_workbook(str(Path(__file__).resolve().parents[2] / "output" / "peer_comparison.xlsx"))
print(f"Sheets: {len(wb.sheetnames)}")
print(wb.sheetnames)