import sqlite3
import sys
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[2]
DB_PATH = str(ROOT / "data" / "nifty100.db")
OUT_PATH = str(ROOT / "output" / "peer_comparison.xlsx")

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GOLD = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

def main():
    conn = sqlite3.connect(DB_PATH)
    pct = pd.read_sql("SELECT * FROM peer_percentiles", conn)
    companies = pd.read_sql("SELECT id AS company_id, company_name FROM companies", conn)
    benchmarks = pd.read_sql("SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn)
    conn.close()

    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for group_name, group_df in pct.groupby("peer_group_name"):
            pivot_val = group_df.pivot(index="company_id", columns="metric", values="value")
            pivot_pct = group_df.pivot(index="company_id", columns="metric", values="percentile_rank")

            pivot_val = pivot_val.merge(companies, on="company_id", how="left").set_index("company_id")
            median_row = pivot_val.median(numeric_only=True)
            pivot_val.loc["MEDIAN"] = median_row

            sheet_name = group_name[:31]
            pivot_val.to_excel(writer, sheet_name=sheet_name)
            ws = writer.sheets[sheet_name]

            bench_ids = set(benchmarks[(benchmarks.peer_group_name == group_name) & (benchmarks.is_benchmark == 1)]["company_id"])

            metric_cols = list(pivot_val.columns)
            for row_idx, cid in enumerate(pivot_val.index, start=2):
                if cid == "MEDIAN":
                    continue
                if cid in bench_ids:
                    for col_idx in range(1, len(metric_cols) + 2):
                        ws.cell(row=row_idx, column=col_idx).fill = GOLD
                    continue
                for col_idx, col in enumerate(metric_cols, start=2):
                    if col not in pivot_pct.columns or cid not in pivot_pct.index:
                        continue
                    p = pivot_pct.loc[cid, col] if col in pivot_pct.columns else None
                    if pd.isna(p):
                        continue
                    fill = GREEN if p >= 0.75 else (RED if p <= 0.25 else YELLOW)
                    ws.cell(row=row_idx, column=col_idx).fill = fill

    print(f"Exported to {OUT_PATH}")

if __name__ == "__main__":
    main()