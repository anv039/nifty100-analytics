import sys
from pathlib import Path
import pandas as pd
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).parent))
from engine import run_preset, PRESETS
from composite_score import compute_composite

ROOT = Path(__file__).resolve().parents[2]
OUT_PATH = str(ROOT / "output" / "screener_output.xlsx")

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

THRESHOLD_COLS = {
    "Quality Compounder": {"return_on_equity_pct": (">", 15), "debt_to_equity": ("<", 1.0)},
    "Value Pick": {"pe_ratio": ("<", 20), "pb_ratio": ("<", 3.0)},
    "Growth Accelerator": {"pat_cagr_5yr": (">", 20), "revenue_cagr_5yr": (">", 15)},
    "Dividend Champion": {"dividend_yield_pct": (">", 2), "dividend_payout": ("<", 80)},
    "Debt-Free Blue Chip": {"debt_to_equity": ("==", 0), "return_on_equity_pct": (">", 12)},
    "Turnaround Watch": {"revenue_cagr_5yr": (">", 10)},
}

def export():
    with pd.ExcelWriter(OUT_PATH, engine="openpyxl") as writer:
        for name in PRESETS:
            df = run_preset(name)
            df = compute_composite(df, sector_relative=True)
            df = df.sort_values("composite_score_final", ascending=False)

            cols = ["company_id", "return_on_equity_pct", "debt_to_equity", "revenue_cagr_5yr",
                    "pat_cagr_5yr", "pe_ratio", "pb_ratio", "dividend_yield_pct", "dividend_payout",
                    "free_cash_flow_cr", "composite_score_final"]
            cols = [c for c in cols if c in df.columns]
            out_df = df[cols]
            sheet_name = name[:31]
            out_df.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]
            thresholds = THRESHOLD_COLS.get(name, {})
            for col_name, (op, val) in thresholds.items():
                if col_name not in cols:
                    continue
                col_idx = cols.index(col_name) + 1
                for row_idx, cell_val in enumerate(out_df[col_name], start=2):
                    if pd.isna(cell_val):
                        continue
                    passed = (op == ">" and cell_val > val) or (op == "<" and cell_val < val) or (op == "==" and cell_val == val)
                    ws.cell(row=row_idx, column=col_idx).fill = GREEN if passed else RED

    print(f"Exported to {OUT_PATH}")

if __name__ == "__main__":
    export()